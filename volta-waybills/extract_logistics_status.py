# -*- coding: utf-8 -*-
"""
Reads "შეკვეთები" sheet from the manually-downloaded Volta Order Management
Google Sheet export and builds a case_id (col A, == CRM Instalment_ID) ->
logistics status (col B) mapping, taking the LATEST row per case_id if a
case_id repeats (order/status can update over time; last row wins).

This is a manual/periodic refresh, same pattern as the other project's
Logistics Daily tab (see volta_logistics_daily.md memory) — no live Google
Sheets API wired up, re-run this whenever a fresh status pull is needed.
"""
import json
import sys
from openpyxl import load_workbook

HERE = r"C:\Users\Lenovo\Desktop\Volta_Waybills"
wb = load_workbook(HERE + r"\volta_order_management.xlsx", data_only=True)
print("sheet names:", wb.sheetnames, file=sys.stderr)

ws = wb["შეკვეთები"]
print(f"dims: {ws.dimensions}, max_row={ws.max_row}", file=sys.stderr)

# print header row + a few data rows for a sanity check
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=3, values_only=True)):
    print(f"row {i+1}: {row[:5]}", file=sys.stderr)

status_by_case = {}
count = 0
skipped = 0
malformed = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    case_id = row[0]
    status = row[1]
    if case_id is None:
        skipped += 1
        continue
    try:
        case_id_int = int(float(case_id))
    except (TypeError, ValueError):
        skipped += 1
        continue
    # real case IDs (Instalment_ID) are 6-digit numbers in the 100000-999999
    # range — a handful of rows in this sheet have a phone number or other
    # unrelated value in column A (data-entry error upstream, not ours to
    # fix), which would otherwise silently create a bogus mapping entry.
    if not (100000 <= case_id_int <= 999999):
        malformed += 1
        continue
    status_by_case[str(case_id_int)] = status
    count += 1

print(f"{count} rows processed, {skipped} skipped (no case id), {malformed} malformed (not a 6-digit case id), {len(status_by_case)} unique case_ids", file=sys.stderr)

with open(HERE + r"\logistics_status_by_case.json", "w", encoding="utf-8") as f:
    json.dump(status_by_case, f, ensure_ascii=False, indent=1, sort_keys=True)

# distribution of statuses, for a sanity check against the Dashboard sheet's own totals
from collections import Counter
dist = Counter(status_by_case.values())
print("status distribution:", file=sys.stderr)
for k, v in dist.most_common():
    print(f"  {k}: {v}", file=sys.stderr)
