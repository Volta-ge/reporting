# -*- coding: utf-8 -*-
"""
Fetches Volta Group's seller-side waybills AND invoices from RS.ge
(2026-01-01 to now), applies validated corrections, and bakes both
datasets into template.html to produce waybill_dashboard.html (which now
covers both the waybills and invoices tabs) in this same folder.

Credentials and field-mapping quirks are documented in the Claude memory
files volta_rsge_api.md / volta_waybill_dashboard.md.
"""
import json
import sys
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path

import pymysql
from lxml import etree
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from zeep import Client

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))
try:
    from config import SU, SP, INVOICE_USER_ID, INVOICE_UN_ID, DB_HOST, DB_PORT, DB_USER, DB_PASS, DB_NAME
except ImportError:
    sys.exit("Missing config.py — copy config.example.py to config.py and fill in real credentials.")

WAYBILL_WSDL = "http://services.rs.ge/WayBillService/WayBillService.asmx?WSDL"
INVOICE_WSDL = "https://www.revenue.mof.ge/ntosservice/ntosservice.asmx?WSDL"
INVOICE_ENDPOINT = "https://www.revenue.mof.ge/ntosservice/ntosservice.asmx"

START = datetime(2026, 1, 1)

RECON_DATE_TOL_DAYS = 45


def fetch_waybills():
    client = Client(WAYBILL_WSDL)
    end = datetime.now()
    result = client.service.get_waybills(
        su=SU, sp=SP,
        itypes=None, buyer_tin=None, statuses=None, car_number=None,
        begin_date_s=START, begin_date_e=end,
        create_date_s=None, create_date_e=None,
        driver_tin=None,
        delivery_date_s=None, delivery_date_e=None,
        full_amount=None, waybill_number=None,
        close_date_s=None, close_date_e=None,
        s_user_ids=None, comment=None,
    )
    waybills = []
    for wb in result.findall("WAYBILL"):
        row = {child.tag: (child.text or "") for child in wb if child.tag not in ("SUB_WAYBILLS", "GOODS_LIST")}
        waybills.append(row)
    return waybills


def build_waybill_rows(raw):
    rows = []
    skipped = 0
    for r in raw:
        if not r.get("WAYBILL_NUMBER"):
            skipped += 1
            continue
        amount = float(r["FULL_AMOUNT"]) if r.get("FULL_AMOUNT") else 0.0
        if r.get("TYPE") == "5":
            amount = -amount
        rows.append({
            "i": r.get("ID", ""),
            "n": r.get("WAYBILL_NUMBER", ""),
            "d": r.get("BEGIN_DATE") or "",
            "b": r.get("BUYER_NAME") or "დაუდგენელი",
            "t": r.get("BUYER_TIN", ""),
            "a": amount,
            "s": r.get("STATUS", ""),
            "y": r.get("TYPE", ""),
            "c": r.get("CAR_NUMBER") or "",
            "e": r.get("END_ADDRESS") or "",
        })
    return rows, skipped


def fetch_invoices():
    """
    get_seller_invoices returns a DataTable as an ADO.NET diffgram whose row
    element is named "invoices" (matches msdata:MainDataTable in the schema)
    under a <DocumentElement> root. zeep's automatic xsd:any binding chokes
    on this shape (LookupError: no element 'DocumentElement'), so the raw
    SOAP envelope is built with zeep and posted/parsed manually instead.
    """
    client = Client(INVOICE_WSDL)
    end = datetime.now()
    node = client.create_message(
        client.service, "get_seller_invoices",
        user_id=INVOICE_USER_ID, un_id=INVOICE_UN_ID,
        s_dt=START, e_dt=end,
        op_s_dt=START, op_e_dt=end,
        invoice_no="", sa_ident_no="", desc="", doc_mos_nom="",
        su=SU, sp=SP,
    )
    envelope = etree.tostring(node)
    headers = {"Content-Type": "text/xml; charset=utf-8", "SOAPAction": "http://tempuri.org/get_seller_invoices"}
    response = client.transport.post(INVOICE_ENDPOINT, envelope, headers)
    root = etree.fromstring(response.content)

    rows = []
    for el in root.iter():
        if etree.QName(el).localname == "invoices":
            row = {etree.QName(child).localname: child.text for child in el}
            rows.append(row)
    return rows


def build_invoice_rows(raw):
    rows = []
    for r in raw:
        f_series = r.get("F_SERIES") or ""
        f_number = r.get("F_NUMBER") or ""
        rows.append({
            "d": r.get("OPERATION_DT") or "",
            "reg": r.get("REG_DT") or "",
            "f": f"{f_series}-{f_number}".strip("-"),
            "b": r.get("ORG_NAME") or "დაუდგენელი",
            "t": r.get("SA_IDENT_NO") or "",
            "a": float(r["TANXA"]) if r.get("TANXA") else 0.0,
            "v": float(r["VAT"]) if r.get("VAT") else 0.0,
            "s": r.get("STATUS") or "",
        })
    return rows


def fetch_crm_sales():
    """
    "Real" CRM sales since START: Order_Status=5 (the business's own
    established real-sale filter, see volta_sales_monthly.md), Product_ID>1
    (excludes placeholder lead rows). Joined to customers for PID (personal
    ID / TIN, the only field shared with RS.ge waybill data) and to users
    for the sales manager's name.
    """
    conn = pymysql.connect(
        host=DB_HOST, port=DB_PORT, user=DB_USER, password=DB_PASS,
        database=DB_NAME, connect_timeout=15,
    )
    cur = conn.cursor(pymysql.cursors.DictCursor)
    cur.execute("""
        SELECT i.Instalment_ID, c.PID, c.FullName, i.Order_Date, i.Full_Cost,
               u.User_FullName AS Manager_Name, p.Model AS Product_Name
        FROM instalments i
        JOIN customers c ON i.Customer_ID = c.Customer_ID
        LEFT JOIN users u ON i.Sales_Manager = u.User_ID
        LEFT JOIN products p ON i.Product_ID = p.Product_ID
        WHERE i.Order_Status = 5
          AND i.Product_ID > 1
          AND i.Order_Date >= %s
          AND c.PID IS NOT NULL AND c.PID <> ''
    """, (START,))
    rows = cur.fetchall()
    conn.close()
    return rows


def amount_tolerance(amount):
    return max(50.0, 0.05 * amount)


import threading
_thread_local = threading.local()


def _thread_client():
    c = getattr(_thread_local, "client", None)
    if c is None:
        c = Client(WAYBILL_WSDL)
        _thread_local.client = c
    return c


def fetch_waybill_goods_names(waybill_id):
    """
    get_waybills (the bulk register call) does not include GOODS_LIST at
    all; only the singular get_waybill(waybill_id) returns it. One zeep
    Client per worker thread (via threading.local) so the WSDL is only
    parsed once per thread, not once per call.
    """
    try:
        client = _thread_client()
        result = client.service.get_waybill(su=SU, sp=SP, waybill_id=int(waybill_id))
        names = [g.findtext("W_NAME") for g in result.findall("GOODS_LIST/GOODS")]
        names = [n for n in names if n]
        return ", ".join(names) if names else ""
    except Exception:
        return ""


def fetch_goods_names_bulk(waybill_ids, max_workers=25):
    """Fetch product names for many waybills concurrently (one SOAP call each — see fetch_waybill_goods_names)."""
    cache = {}
    ids = list(dict.fromkeys(waybill_ids))
    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        futures = {ex.submit(fetch_waybill_goods_names, wid): wid for wid in ids}
        for fut in as_completed(futures):
            wid = futures[fut]
            try:
                cache[wid] = fut.result()
            except Exception:
                cache[wid] = ""
    return cache


def build_reconciliation(crm_rows, waybill_raw):
    """
    Per-PID bipartite match: each CRM sale is paired with at most one
    waybill on the same buyer TIN, within RECON_DATE_TOL_DAYS days and a
    5%-or-50-GEL amount tolerance, greedily assigning the closest pairs
    first (by combined normalized date+amount distance) so one waybill
    can't double-count against two different sales. Cancelled (-2) and
    return (TYPE=5) waybills are excluded from the matching pool — they
    don't represent a valid "goods delivered for this sale" event.

    Returns three views (embedded together so the dashboard tab can filter
    client-side without re-fetching):
    - "matched": one row per matched CRM sale, with its actual paired waybill.
    - "missing": one row per unmatched CRM sale, with the nearest same-TIN
      waybill (regardless of tolerance) as a diagnostic, since the amount
      tolerance can legitimately miss cases where one CRM sale's goods were
      delivered as two separate partial waybills (see "byPerson" below).
    - "byPerson": every CRM sale and every waybill for every buyer TIN that
      has at least one CRM sale in the period, grouped by TIN and sorted by
      date, so a person's full timeline can be eyeballed manually — this is
      the fallback for exactly the split-delivery case the amount-tolerance
      matching can't handle on its own (one sale, two partial waybills).
    """
    valid_wb = [
        w for w in waybill_raw
        if w.get("WAYBILL_NUMBER") and w.get("STATUS") != "-2" and w.get("TYPE") != "5" and w.get("BEGIN_DATE")
    ]
    # Permissive sibling of valid_wb: same filters minus the TYPE=5 (return)
    # exclusion, amount pre-negated for returns to match the main waybills
    # tab's own sign convention (see build_waybill_rows). Used ONLY for the
    # byPerson display/total — a return still can't be a *match candidate*
    # for a new CRM sale (valid_wb/wb_by_tin stay unchanged for that), but
    # hiding it entirely from a person's card was a real bug: a delivery
    # that was later returned looks identical to "staff never issued a
    # waybill" if the return itself is invisible. Found 2026-08-27 via PID
    # 13001011126 — an iPhone 16 waybill (+4,060) had no CRM sale behind it
    # (flagged "აკლია გაყიდვა"), but RS.ge also shows a -4,060 return for the
    # same buyer that never appeared anywhere in the tab; including it flips
    # that person's netDiff to ~0 (correctly matched, nothing missing).
    all_wb = [
        w for w in waybill_raw
        if w.get("WAYBILL_NUMBER") and w.get("STATUS") != "-2" and w.get("BEGIN_DATE")
    ]
    crm_by_pid = defaultdict(list)
    for r in crm_rows:
        crm_by_pid[r["PID"]].append(r)
    wb_by_tin = defaultdict(list)
    for w in valid_wb:
        wb_by_tin[w["BUYER_TIN"]].append(w)
    wb_by_tin_all = defaultdict(list)
    for w in all_wb:
        wb_by_tin_all[w["BUYER_TIN"]].append(w)

    # Product names require one SOAP call per waybill (get_waybills, the bulk
    # register call, has no GOODS_LIST at all) — prefetch concurrently for
    # every waybill belonging to a buyer who has at least one CRM sale in the
    # period (matching pool AND the fuller display list, so returns get a
    # product name too), so matched/missing/byPerson can all reuse the cache.
    relevant_wb_ids = [w["ID"] for pid in crm_by_pid for w in wb_by_tin_all.get(pid, [])]
    print(f"  prefetching product names for {len(set(relevant_wb_ids))} waybills...", file=sys.stderr)
    goods_cache = fetch_goods_names_bulk(relevant_wb_ids)

    matched = []
    missing = []
    by_person = []

    for pid, sales in crm_by_pid.items():
        candidates = wb_by_tin.get(pid, [])
        pairs = []
        for ci, c in enumerate(sales):
            c_amt = float(c["Full_Cost"])
            c_date = c["Order_Date"]
            tol = amount_tolerance(c_amt)
            for wi, w in enumerate(candidates):
                w_amt = float(w["FULL_AMOUNT"]) if w.get("FULL_AMOUNT") else 0.0
                w_date = datetime.fromisoformat(w["BEGIN_DATE"])
                date_diff = abs((w_date - c_date).days)
                amt_diff = abs(w_amt - c_amt)
                if date_diff <= RECON_DATE_TOL_DAYS and amt_diff <= tol:
                    score = date_diff / RECON_DATE_TOL_DAYS + amt_diff / tol
                    pairs.append((score, ci, wi))
        pairs.sort(key=lambda x: x[0])
        used_ci, used_wi = set(), set()
        for score, ci, wi in pairs:
            if ci in used_ci or wi in used_wi:
                continue
            used_ci.add(ci)
            used_wi.add(wi)
            c = sales[ci]
            w = candidates[wi]
            c_amt = float(c["Full_Cost"])
            w_amt = float(w["FULL_AMOUNT"]) if w.get("FULL_AMOUNT") else 0.0
            matched.append({
                "d": c["Order_Date"].isoformat(),
                "id": c["Instalment_ID"],
                "b": c["FullName"] or "",
                "t": pid,
                "mgr": c.get("Manager_Name") or "—",
                "prodDb": c.get("Product_Name") or "—",
                "amtDb": c_amt,
                "prodWb": goods_cache.get(w["ID"], "") or "—",
                "amtWb": w_amt,
                "dateWb": w["BEGIN_DATE"],
                "wbNumber": w.get("WAYBILL_NUMBER", ""),
                "diff": w_amt - c_amt,
            })

        for ci, c in enumerate(sales):
            if ci in used_ci:
                continue
            c_amt = float(c["Full_Cost"])
            c_date = c["Order_Date"]
            if not candidates:
                wb_prod, wb_amt, wb_date, wb_number = None, None, None, None
            else:
                nearest = min(candidates, key=lambda w: abs((datetime.fromisoformat(w["BEGIN_DATE"]) - c_date).days))
                wb_number = nearest.get("WAYBILL_NUMBER", "")
                wb_date = nearest["BEGIN_DATE"]
                wb_amt = float(nearest["FULL_AMOUNT"]) if nearest.get("FULL_AMOUNT") else 0.0
                wb_prod = goods_cache.get(nearest["ID"], "")
            missing.append({
                "d": c_date.isoformat(),
                "id": c["Instalment_ID"],
                "b": c["FullName"] or "",
                "t": pid,
                "mgr": c.get("Manager_Name") or "—",
                "prodDb": c.get("Product_Name") or "—",
                "amtDb": c_amt,
                "prodWb": wb_prod or "—",
                "amtWb": wb_amt,
                "dateWb": wb_date,
                "wbNumber": wb_number,
                "diff": (wb_amt - c_amt) if wb_amt is not None else None,
            })

        matched_wb_ids = {candidates[wi]["ID"] for wi in used_wi}
        all_candidates = sorted(wb_by_tin_all.get(pid, []), key=lambda w: w["BEGIN_DATE"])
        name = sales[0]["FullName"] or ""
        by_person.append({
            "t": pid,
            "b": name,
            "sales": [
                {
                    "d": c["Order_Date"].isoformat(),
                    "id": c["Instalment_ID"],
                    "prod": c.get("Product_Name") or "—",
                    "a": float(c["Full_Cost"]),
                    "mgr": c.get("Manager_Name") or "—",
                    "matched": ci in used_ci,
                }
                for ci, c in enumerate(sales)
            ],
            "waybills": [
                {
                    "d": w["BEGIN_DATE"],
                    "n": w.get("WAYBILL_NUMBER", ""),
                    "prod": goods_cache.get(w["ID"], "") or "—",
                    "a": (-float(w["FULL_AMOUNT"]) if w.get("TYPE") == "5" else float(w["FULL_AMOUNT"])) if w.get("FULL_AMOUNT") else 0.0,
                    "matched": w["ID"] in matched_wb_ids,
                    "isReturn": w.get("TYPE") == "5",
                }
                for w in all_candidates
            ],
        })

    # Person-level net-difference: the trustworthy "is something actually
    # wrong here?" signal, matching the dashboard's computePeopleForRange()
    # logic exactly (see volta_sales_waybill_reconciliation.md). A person's
    # individual sale<->waybill pairwise match can fail the ±45d/±5%
    # tolerance while their totals still net to ~0 (goods delivered via
    # split/partial waybills) — that's not a real gap, so status is judged
    # by (CRM total - waybill total) per PID, never by counting sale-level
    # match flags. The SIGN of that difference tells us which of two very
    # different problems it is: positive means a real CRM sale has no
    # waybill ("აკლია ზედნადები" — compliance risk); negative means a
    # waybill exists for goods that were never recorded as a CRM sale
    # ("აკლია გაყიდვა" — an unrecorded sale). Lumping both into one
    # "missing" bucket hid the second failure mode entirely (see the
    # "positive vs negative net-diff" fix note).
    NET_TOL = 1.0
    for p in by_person:
        sales_total = sum(s["a"] for s in p["sales"])
        wb_total = sum(w["a"] for w in p["waybills"])
        p["salesTotal"] = sales_total
        p["wbTotal"] = wb_total
        p["netDiff"] = sales_total - wb_total

    missing_wb_people = [p for p in by_person if p["netDiff"] >= NET_TOL]
    missing_sale_people = [p for p in by_person if p["netDiff"] <= -NET_TOL]
    matched_people = [p for p in by_person if abs(p["netDiff"]) < NET_TOL]
    risk_amount_wb = sum(p["netDiff"] for p in missing_wb_people)
    risk_amount_sale = sum(-p["netDiff"] for p in missing_sale_people)

    total = len(crm_rows)
    missing_amount = sum(m["amtDb"] for m in missing)
    return {
        "summary": {
            # sale-level flat counts — kept for reference only, NOT the
            # number to show/export as "missing". Use the people.* fields.
            "total": total,
            "matched": len(matched),
            "missing": len(missing),
            "missingAmount": missing_amount,
            # person-level (net-diff) counts — these are the trustworthy numbers.
            "totalPeople": len(by_person),
            "matchedPeople": len(matched_people),
            "missingWbPeople": len(missing_wb_people),
            "missingSalePeople": len(missing_sale_people),
            "riskAmountWb": risk_amount_wb,
            "riskAmountSale": risk_amount_sale,
            "riskAmount": risk_amount_wb + risk_amount_sale,
        },
        "matched": matched,
        "missing": missing,
        "byPerson": by_person,
    }


def build_excel_report(recon, wb_rows, inv_rows):
    """
    Full "Accounting" export (fixed filename, overwritten daily) — every tab
    of the merged Accounting nav-group (see volta_analytics_merge memory),
    not just reconciliation: the raw ზედნადებები/ანგარიშ-ფაქტურები registers
    plus the full person-level reconciliation. Mirrors everything the
    dashboard holds so the underlying data can be audited/filtered/corrected
    directly in Excel rather than only spot-checked in the browser. A person
    counts as "აკლია ზედნადები" only if their CRM-sales total and waybill
    total don't net out (|diff| >= 1 GEL), never by a raw per-sale
    match-flag count — see the "KPI/chart consistency fix" note in
    volta_sales_waybill_reconciliation.md. Six sheets: ზედნადებები (full
    waybill register), ანგარიშ-ფაქტურები (full invoice register), შეჯამება
    (reconciliation summary), პირები (one row per buyer — ALL of them,
    matched and missing alike, with a status column + AutoFilter), დეტალები
    (every sale + waybill line item for every one of those buyers),
    გაყიდვები პირადობის მიხედვით (person-card format).
    """
    summary = recon["summary"]
    by_person = recon["byPerson"]
    NET_TOL = 1.0
    all_people = sorted(by_person, key=lambda p: abs(p["netDiff"]), reverse=True)

    def person_status(p):
        if p["netDiff"] >= NET_TOL:
            return "აკლია ზედნადები"
        if p["netDiff"] <= -NET_TOL:
            return "აკლია გაყიდვა"
        return "დაწყვილებული"

    base_font = "Arial"
    header_fill = PatternFill(start_color="1B2224", end_color="1B2224", fill_type="solid")
    header_font = Font(name=base_font, bold=True, color="FFFFFF", size=11)
    title_font = Font(name=base_font, bold=True, size=14)
    note_font = Font(name=base_font, italic=True, size=9, color="5B6B6E")
    cell_font = Font(name=base_font, size=10)
    bold_font = Font(name=base_font, bold=True, size=10)

    wb = Workbook()

    STATUS_MAP_WB = {
        "2": "დასრულებული", "1": "აქტიური", "8": "გადამზიდავთან",
        "-2": "გაუქმებული", "0": "შენახული", "-1": "წაშლილი",
    }
    TYPE_MAP_WB = {
        "1": "შიდა გადაზიდვა", "2": "მიწოდება ტრანსპორტირებით", "3": "მიწოდება ტრანსპორტირების გარეშე",
        "4": "დისტრიბუცია", "5": "საქონლის დაბრუნება", "6": "ქვე-ზედნადები",
    }
    STATUS_MAP_INV = {
        "0": "ახალი", "1": "დასადასტურებელი", "2": "დადასტურებული", "3": "კორექტირებული (პირველადი)",
        "4": "ახალი კორექტირება", "5": "კორექტ. დასადასტურებელი", "8": "კორექტ. დადასტურებული",
    }

    def naive_dt(s):
        # REG_DT (invoices) comes back with a timezone offset; BEGIN_DATE/OPERATION_DT
        # don't. openpyxl refuses to write a tz-aware datetime at all, so always strip it.
        if not s:
            return None
        return datetime.fromisoformat(s).replace(tzinfo=None)

    ws0 = wb.active
    ws0.title = "ზედნადებები"
    headers0 = ["თარიღი", "ზედნადების №", "მყიდველი", "პ/ნ", "თანხა (₾)", "სტატუსი", "ტიპი", "ავტომობილი", "მისამართი"]
    for col, h in enumerate(headers0, start=1):
        c = ws0.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, r in enumerate(sorted(wb_rows, key=lambda r: r["d"], reverse=True), start=2):
        ws0.cell(row=i, column=1, value=naive_dt(r["d"])).number_format = "DD.MM.YYYY HH:MM"
        ws0.cell(row=i, column=2, value=r["n"])
        ws0.cell(row=i, column=3, value=r["b"])
        ws0.cell(row=i, column=4, value=r["t"])
        ws0.cell(row=i, column=5, value=r["a"]).number_format = "#,##0"
        ws0.cell(row=i, column=6, value=STATUS_MAP_WB.get(r["s"], r["s"]))
        ws0.cell(row=i, column=7, value=TYPE_MAP_WB.get(r["y"], r["y"]))
        ws0.cell(row=i, column=8, value=r["c"])
        ws0.cell(row=i, column=9, value=r["e"])
        for col in range(1, 10):
            ws0.cell(row=i, column=col).font = cell_font
    last_row0 = len(wb_rows) + 1
    widths0 = [18, 14, 34, 13, 12, 14, 24, 12, 40]
    for col, w in enumerate(widths0, start=1):
        ws0.column_dimensions[get_column_letter(col)].width = w
    ws0.freeze_panes = "A2"
    ws0.auto_filter.ref = f"A1:I{last_row0}"

    ws00 = wb.create_sheet("ანგარიშ-ფაქტურები")
    headers00 = ["პერიოდი", "რეგისტრაციის თარიღი", "ფაქტურის №", "მყიდველი", "პ/ნ", "თანხა (₾)", "დღგ (₾)", "სტატუსი"]
    for col, h in enumerate(headers00, start=1):
        c = ws00.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, r in enumerate(sorted(inv_rows, key=lambda r: r["d"], reverse=True), start=2):
        ws00.cell(row=i, column=1, value=naive_dt(r["d"])).number_format = "DD.MM.YYYY"
        ws00.cell(row=i, column=2, value=naive_dt(r["reg"])).number_format = "DD.MM.YYYY HH:MM"
        ws00.cell(row=i, column=3, value=r["f"])
        ws00.cell(row=i, column=4, value=r["b"])
        ws00.cell(row=i, column=5, value=r["t"])
        ws00.cell(row=i, column=6, value=r["a"]).number_format = "#,##0"
        ws00.cell(row=i, column=7, value=r["v"]).number_format = "#,##0"
        ws00.cell(row=i, column=8, value=STATUS_MAP_INV.get(r["s"], r["s"]))
        for col in range(1, 9):
            ws00.cell(row=i, column=col).font = cell_font
    last_row00 = len(inv_rows) + 1
    widths00 = [14, 18, 14, 34, 13, 12, 12, 22]
    for col, w in enumerate(widths00, start=1):
        ws00.column_dimensions[get_column_letter(col)].width = w
    ws00.freeze_panes = "A2"
    ws00.auto_filter.ref = f"A1:H{last_row00}"

    ws = wb.create_sheet("შეჯამება")
    ws["A1"] = "ვოლტა — გაყიდვა ↔ ზედნადები შედარება"
    ws["A1"].font = title_font
    ws["A2"] = ("წყარო: myvolta.info CRM (instalments, Order_Status=5) + RS.ge WayBillService · "
                "გენერირებულია: " + datetime.now().strftime("%d.%m.%Y %H:%M"))
    ws["A2"].font = note_font

    labels = [
        ("სულ მყიდველი (2026-01-01-დან, PID-ით)", summary["totalPeople"]),
        ("დაწყვილებული მყიდველი (სხვაობა < 1₾)", summary["matchedPeople"]),
        ("აკლია ზედნადები (მყიდველი — გაყიდვაა, ზედნადები არა)", summary["missingWbPeople"]),
        ("რისკის ქვეშ თანხა — აკლია ზედნადები (₾)", summary["riskAmountWb"]),
        ("აკლია გაყიდვა (მყიდველი — ზედნადებია, გაყიდვა CRM-ში არა)", summary["missingSalePeople"]),
        ("რისკის ქვეშ თანხა — აკლია გაყიდვა (₾)", summary["riskAmountSale"]),
    ]
    row = 4
    for label, val in labels:
        ws.cell(row=row, column=1, value=label).font = cell_font
        c = ws.cell(row=row, column=2, value=val)
        c.font = bold_font
        if "თანხა" in label:
            c.number_format = "#,##0"
        row += 1

    ws.cell(row=row, column=1, value="დაფარვის % (=B5/B4)").font = cell_font
    cov = ws.cell(row=row, column=2, value="=B5/B4")
    cov.font = bold_font
    cov.number_format = "0.0%"
    row += 2

    ws.cell(row=row, column=1, value="მეთოდოლოგია:").font = bold_font
    row += 1
    methodology = [
        "დაშვება არის პირზე (პ/ნ = TIN) დაფუძნებული: შედარებულია მისი CRM გაყიდვების ჯამური თანხა და",
        "ამავე პერიოდის ზედნადებების ჯამური თანხა. თუ სხვაობა 1₾-ზე ნაკლებია, პირი ითვლება დაწყვილებულად —",
        "თუნდაც ცალკეული გაყიდვა-ზედნადები წყვილმა ვერ გაიაროს ±45-დღიანი/±5%-იანი ტოლერანტობა",
        "(მაგ. ერთი გაყიდვა მიტანილია ორი ცალკეული ზედნადებით). სხვაობის ნიშანი განსაზღვრავს პრობლემის ტიპს:",
        "დადებითი სხვაობა (CRM > ზედნადები) ნიშნავს 'აკლია ზედნადები'-ს — რეალური გაყიდვაა გატარებული,",
        "მაგრამ ზედნადები არ გამოწერილა (საგადასახადო რისკი). უარყოფითი სხვაობა (ზედნადები > CRM) ნიშნავს",
        "'აკლია გაყიდვა'-ს — ზედნადები გამოწერილია, მაგრამ CRM-ში შესაბამისი გაყიდვა არ ჩანს (შემოსავალი",
        "შეიძლება არასწორად აღრიცხულა). 'პირები' გვერდზე ყველა მყიდველია — გამოიყენე 'სტატუსი' სვეტის",
        "ფილტრი (AutoFilter, C სვეტი) ბრაუზერის ყველა/დაწყვილებული/აკლია ზედნადები/აკლია გაყიდვა ფილტრების",
        "იმეორებლად. 'დეტალები' გვერდზე ყველა ამ მყიდველის ცალკეული გაყიდვა და ზედნადები ცალ-ცალკეა —",
        "ეს არის ზუსტად ის, რასაც დეშბორდზე პირის ბარათი აჩვენებს.",
        "გაუქმებული ზედნადებები არცერთ ჯამში არ ითვლება. 'უკან დაბრუნება' ტიპის ზედნადებები",
        "ჩართულია პირის ჯამში (თანხა უარყოფითად, RS.ge-ის საკუთარი კონვენციის მიხედვით) — ისინი",
        "მხოლოდ ახალი გაყიდვის დამწყვილებელ კანდიდატებად არ განიხილება, დაბრუნება ხომ არ ადასტურებს",
        "ახალ მიწოდებას; მაგრამ პირის საერთო სხვაობის (და სტატუსის) გამოთვლაში სრულად ჩართულია.",
    ]
    for line in methodology:
        ws.cell(row=row, column=1, value=line).font = note_font
        row += 1

    ws.column_dimensions["A"].width = 62
    ws.column_dimensions["B"].width = 16

    ws2 = wb.create_sheet("პირები")
    headers = [
        "მყიდველი", "პ/ნ", "სტატუსი", "CRM ჯამი (₾)", "ზედნადებების ჯამი (₾)", "სხვაობა (₾)",
        "გაყიდვების რაოდ.", "ზედნადებების რაოდ.",
    ]
    for col, h in enumerate(headers, start=1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, p in enumerate(all_people, start=2):
        status = person_status(p)
        ws2.cell(row=i, column=1, value=p["b"])
        ws2.cell(row=i, column=2, value=p["t"])
        ws2.cell(row=i, column=3, value=status)
        ws2.cell(row=i, column=4, value=p["salesTotal"]).number_format = "#,##0"
        ws2.cell(row=i, column=5, value=p["wbTotal"]).number_format = "#,##0"
        ws2.cell(row=i, column=6, value=p["netDiff"]).number_format = "#,##0;(#,##0)"
        ws2.cell(row=i, column=7, value=len(p["sales"]))
        ws2.cell(row=i, column=8, value=len(p["waybills"]))
        for col in range(1, 9):
            ws2.cell(row=i, column=col).font = cell_font

    last_row = len(all_people) + 1
    total_row = last_row + 2
    ws2.cell(row=total_row, column=1, value="სულ:").font = bold_font
    for col, letter in ((4, "D"), (5, "E"), (6, "F")):
        tc = ws2.cell(row=total_row, column=col, value=f"=SUM({letter}2:{letter}{last_row})")
        tc.font = bold_font
        tc.number_format = "#,##0"

    widths = [26, 14, 18, 16, 18, 14, 14, 16]
    for col, w in enumerate(widths, start=1):
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:H{last_row}"

    ws3 = wb.create_sheet("დეტალები")
    headers3 = [
        "მყიდველი", "პ/ნ", "სტატუსი", "პირის სხვაობა (₾)", "ტიპი", "თარიღი",
        "საბუთის №", "პროდუქტი", "თანხა (₾)", "დაწყვილებული",
    ]
    for col, h in enumerate(headers3, start=1):
        c = ws3.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    r = 2
    for p in all_people:
        status = person_status(p)
        items = (
            [("გაყიდვა", s["d"], s["id"], s["prod"], s["a"], s["matched"]) for s in p["sales"]]
            + [("დაბრუნება" if w.get("isReturn") else "ზედნადები", w["d"], w["n"], w["prod"], w["a"], w["matched"]) for w in p["waybills"]]
        )
        items.sort(key=lambda x: x[1])
        for kind, d, doc_no, prod, amt, matched in items:
            ws3.cell(row=r, column=1, value=p["b"])
            ws3.cell(row=r, column=2, value=p["t"])
            ws3.cell(row=r, column=3, value=status)
            ws3.cell(row=r, column=4, value=p["netDiff"]).number_format = "#,##0;(#,##0)"
            ws3.cell(row=r, column=5, value=kind)
            ws3.cell(row=r, column=6, value=datetime.fromisoformat(d)).number_format = "DD.MM.YYYY HH:MM"
            ws3.cell(row=r, column=7, value=doc_no)
            ws3.cell(row=r, column=8, value=prod)
            ws3.cell(row=r, column=9, value=amt).number_format = "#,##0"
            ws3.cell(row=r, column=10, value="კი" if matched else "არა")
            for col in range(1, 11):
                ws3.cell(row=r, column=col).font = cell_font
            r += 1

    last_row3 = r - 1
    widths3 = [26, 14, 18, 16, 12, 16, 14, 34, 14, 14]
    for col, w in enumerate(widths3, start=1):
        ws3.column_dimensions[get_column_letter(col)].width = w
    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = f"A1:J{last_row3}"

    # ---- "გაყიდვები პირადობის მიხედვით" — the dashboard's person-card view,
    # reproduced block-by-block: CRM sales and waybills side by side per
    # person, in the same newest-sale-first order as the dashboard.
    green = Font(name=base_font, size=10, color="1F7A3D")
    red = Font(name=base_font, size=10, color="B23A3A")
    gray_font = Font(name=base_font, size=10, color="808080")
    amber_font = Font(name=base_font, size=10, color="B8631F")
    empty_font = Font(name=base_font, italic=True, size=9, color="8A9598")
    group_header_font = Font(name=base_font, bold=True, size=9, color="5B6B6E")
    status_color = {"დაწყვილებული": "1F7A3D", "აკლია ზედნადები": "B23A3A", "აკლია გაყიდვა": "B08A16"}

    ws4 = wb.create_sheet("გაყიდვები პირადობის მიხედვით")
    ws4["A1"] = "ვოლტა — გაყიდვები პირადობის მიხედვით (დეშბორდის ბარათების ფორმატი)"
    ws4["A1"].font = title_font
    ws4["A2"] = "დალაგებულია მყიდველის უახლესი CRM გაყიდვის თარიღით (ჯერ უახლესი) — ისევე, როგორც დეშბორდზე."
    ws4["A2"].font = note_font

    def latest_sale_date(p):
        return max((s["d"] for s in p["sales"]), default="")

    people_by_recency = sorted(all_people, key=latest_sale_date, reverse=True)

    r = 4
    for p in people_by_recency:
        status = person_status(p)
        sc = status_color[status]
        head = (f'{p["b"]}  ·  პ/ნ {p["t"]}  ·  CRM {p["salesTotal"]:,.0f}₾ / '
                f'ზედნადები {p["wbTotal"]:,.0f}₾  ·  {status}')
        ws4.cell(row=r, column=1, value=head).font = Font(name=base_font, bold=True, size=11, color=sc)
        dc = ws4.cell(row=r, column=7, value=p["netDiff"])
        dc.number_format = "#,##0;(#,##0)"
        dc.font = Font(name=base_font, bold=True, size=10, color=sc)
        r += 1

        sales_sorted = sorted(p["sales"], key=lambda s: s["d"])
        wbs_sorted = sorted(p["waybills"], key=lambda w: w["d"])
        ws4.cell(row=r, column=1, value="თარიღი").font = group_header_font
        ws4.cell(row=r, column=2, value=f"CRM გაყიდვები ({len(sales_sorted)})").font = group_header_font
        ws4.cell(row=r, column=3, value="თანხა (₾)").font = group_header_font
        ws4.cell(row=r, column=5, value="თარიღი").font = group_header_font
        ws4.cell(row=r, column=6, value=f"ზედნადებები ({len(wbs_sorted)})").font = group_header_font
        ws4.cell(row=r, column=7, value="თანხა (₾)").font = group_header_font
        r += 1

        n = max(len(sales_sorted), len(wbs_sorted), 1)
        for i in range(n):
            if i < len(sales_sorted):
                s = sales_sorted[i]
                fnt = green if s["matched"] else red
                ws4.cell(row=r, column=1, value=datetime.fromisoformat(s["d"])).number_format = "DD.MM.YYYY"
                ws4.cell(row=r, column=1).font = fnt
                ws4.cell(row=r, column=2, value=s["prod"]).font = fnt
                c = ws4.cell(row=r, column=3, value=s["a"])
                c.number_format = "#,##0"
                c.font = fnt
            elif i == 0:
                ws4.cell(row=r, column=2, value="ამ პერიოდში გაყიდვა არ არის").font = empty_font
            if i < len(wbs_sorted):
                w = wbs_sorted[i]
                fnt = amber_font if w.get("isReturn") else (green if w["matched"] else gray_font)
                prod_label = w["prod"] + (" · დაბრუნება" if w.get("isReturn") else "")
                ws4.cell(row=r, column=5, value=datetime.fromisoformat(w["d"])).number_format = "DD.MM.YYYY"
                ws4.cell(row=r, column=5).font = fnt
                ws4.cell(row=r, column=6, value=prod_label).font = fnt
                c = ws4.cell(row=r, column=7, value=w["a"])
                c.number_format = "#,##0"
                c.font = fnt
            elif i == 0:
                ws4.cell(row=r, column=6, value="ამ პერიოდში ზედნადები არ არის").font = empty_font
            r += 1
        r += 1  # spacer row between person blocks

    widths4 = [14, 36, 12, 3, 14, 36, 12]
    for col, w in enumerate(widths4, start=1):
        ws4.column_dimensions[get_column_letter(col)].width = w

    out_path = HERE / "Volta_აკლია_ზედნადები.xlsx"
    wb.save(out_path)
    return out_path


def main():
    print(f"[{datetime.now()}] fetching waybills from RS.ge...", file=sys.stderr)
    raw_wb = fetch_waybills()
    wb_rows, skipped = build_waybill_rows(raw_wb)
    print(f"waybills: fetched {len(raw_wb)}, kept {len(wb_rows)}, skipped {skipped} unnumbered", file=sys.stderr)

    print(f"[{datetime.now()}] fetching invoices from RS.ge...", file=sys.stderr)
    raw_inv = fetch_invoices()
    inv_rows = build_invoice_rows(raw_inv)
    print(f"invoices: fetched {len(inv_rows)}", file=sys.stderr)

    print(f"[{datetime.now()}] fetching CRM sales from MySQL...", file=sys.stderr)
    crm_rows = fetch_crm_sales()
    print(f"CRM sales (Order_Status=5, Product_ID>1): {len(crm_rows)}", file=sys.stderr)

    recon = build_reconciliation(crm_rows, raw_wb)
    s = recon["summary"]
    print(f"reconciliation (person-level): matched {s['matchedPeople']}, "
          f"missing waybill {s['missingWbPeople']} people ({s['riskAmountWb']:.0f} GEL), "
          f"missing sale {s['missingSalePeople']} people ({s['riskAmountSale']:.0f} GEL)", file=sys.stderr)

    xlsx_path = build_excel_report(recon, wb_rows, inv_rows)
    print(f"wrote {xlsx_path}", file=sys.stderr)

    wb_json = json.dumps(wb_rows, ensure_ascii=False, separators=(",", ":"))
    inv_json = json.dumps(inv_rows, ensure_ascii=False, separators=(",", ":"))
    recon_json = json.dumps(recon, ensure_ascii=False, separators=(",", ":"), default=str)

    template = (HERE / "template.html").read_text(encoding="utf-8")
    out_html = (template
                .replace("__DATA_WB__", wb_json)
                .replace("__DATA_INV__", inv_json)
                .replace("__DATA_RECON__", recon_json))
    out_path = HERE / "waybill_dashboard.html"
    out_path.write_text(out_html, encoding="utf-8")
    print(f"wrote {out_path} ({len(out_html.encode('utf-8'))} bytes)", file=sys.stderr)


if __name__ == "__main__":
    main()
